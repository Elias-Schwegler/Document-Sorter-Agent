import os
import logging
from pathlib import Path

logger = logging.getLogger(__name__)


def extract_text(filepath: str, lang: str = "eng") -> tuple[str, int]:
    """Extract text from a document file.

    Returns (text, page_count). On failure returns ("", 0).
    """
    try:
        ext = Path(filepath).suffix.lower().lstrip(".")
        if ext == "pdf":
            return _extract_pdf(filepath, lang)
        elif ext in ("png", "jpg", "jpeg", "tiff", "tif", "bmp", "gif", "webp"):
            return _extract_image(filepath, lang)
        elif ext == "docx":
            return _extract_docx(filepath)
        elif ext == "xlsx":
            return _extract_xlsx(filepath)
        elif ext in ("txt", "md", "csv", "rtf"):
            return _extract_plaintext(filepath)
        else:
            logger.warning("Unsupported file type: %s", ext)
            return "", 0
    except Exception as e:
        logger.error("Failed to extract text from %s: %s", filepath, e)
        return "", 0


def _extract_pdf(filepath: str, lang: str) -> tuple[str, int]:
    import fitz  # pymupdf

    doc = fitz.open(filepath)
    page_count = len(doc)
    pages_text: list[str] = []

    for page in doc:
        text = page.get_text().strip()
        if text:
            pages_text.append(text)
        else:
            # Fall back to OCR for scanned pages
            ocr_text = _ocr_page_image(page, lang)
            if ocr_text:
                pages_text.append(ocr_text)

    doc.close()
    return "\n\n".join(pages_text), page_count


def _ocr_page_image(page, lang: str) -> str:
    """Render a PDF page to an image and OCR it."""
    try:
        import pytesseract
        from PIL import Image
        import io

        pix = page.get_pixmap(dpi=300)
        img = Image.open(io.BytesIO(pix.tobytes("png")))
        text = pytesseract.image_to_string(img, lang=lang).strip()
        return text
    except Exception as e:
        logger.warning("OCR fallback failed for page: %s", e)
        return ""


def _extract_image(filepath: str, lang: str) -> tuple[str, int]:
    import pytesseract
    from PIL import Image

    img = Image.open(filepath)
    text = pytesseract.image_to_string(img, lang=lang).strip()
    return text, 1


def _extract_docx(filepath: str) -> tuple[str, int]:
    from docx import Document

    doc = Document(filepath)
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    text = "\n\n".join(paragraphs)
    # Approximate page count from paragraph count
    page_count = max(1, len(paragraphs) // 25)
    return text, page_count


def _extract_xlsx(filepath: str) -> tuple[str, int]:
    from openpyxl import load_workbook

    wb = load_workbook(filepath, read_only=True, data_only=True)
    sheets_text: list[str] = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            row_text = "\t".join(cells).strip()
            if row_text:
                rows.append(row_text)
        if rows:
            sheets_text.append(f"[Sheet: {sheet}]\n" + "\n".join(rows))

    wb.close()
    text = "\n\n".join(sheets_text)
    page_count = len(wb.sheetnames) if sheets_text else 1
    return text, page_count


def _extract_plaintext(filepath: str) -> tuple[str, int]:
    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        text = f.read()
    # Approximate page count (roughly 3000 chars per page)
    page_count = max(1, len(text) // 3000)
    return text, page_count
